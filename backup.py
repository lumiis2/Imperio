import customtkinter
from tkinter import filedialog
import pandas as pd
import subprocess   
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

def set_var(self):
        self.root = None
        self.tolerance = 0.5
        self.var_name = "CASTELO"
        self.csv_data = pd.DataFrame()
        self.excel_data = pd.DataFrame()
        self.general_data = pd.DataFrame()

        ### QUERO DIMINUIR ESSAS VARS GLOBAIS
        self.rede_storage =[]
        self.w3_storage = []
        self.w3_storage_s = []
        self.rede_storage_s = []
        self.pares_encontrados = []
        self.csv_path = " "
        self.excel_path = " "
        self.temp_excel_path = " "
        self.temp_csv_path = " "
        

def choose_store(self, store):
        store_mappings = {
            "Loja Castelo": "CASTELO",
            "Loja Cidade Nova": "CID. NOVA",
            "Loja Planalto": "PLANALTO",
            "Loja Contagem": "CONTAGEM",
            "Loja Nova Lima": "NOVA LIMA",
            "Loja E-commerce": "E-COMM"
        }
        self.var_name = store_mappings.get(store, "Loja não encontrada")

def download(self, file_type):
    file_path = filedialog.askopenfilename(
        title=f"Select the {file_type.upper()} file",
        filetypes=[(f"{file_type.upper()} files", f"*.{file_type}")]
    )
    if file_path:
        if file_type == 'csv':
            self.csv_data = csv_read(self, file_path)
            self.csv_path = file_path
        elif file_type == 'xlsx':
            self.excel_data = excel_read(self, file_path)
            self.excel_path = file_path

        create_compare_button(self)

def create_compare_button(self):
    if not self.csv_data.empty and not self.excel_data.empty:
        self.button_3 = customtkinter.CTkButton(self, text="Comparar", command=lambda: comparer(self))
        self.button_3.grid(row=3, column=0, padx=20, pady=10)

def comparer(self):
        general = False
        if self.temp_excel_path == self.excel_path and self.temp_csv_path == self.csv_path:
            general = True
            process(self, 3, 2, general)
            open_sheets(self)
        else:
            process(self, 'Total', 2, general)
            open_sheets(self)
        self.temp_excel_path = self.excel_path
        self.temp_csv_path = self.csv_path

def excel_read(self, file_path, verbose=False):
        compare_col = 2
        var_skip = 0
        count = 0
        ind = 0

        try:
            self.excel_data = pd.read_excel(file_path, header=None)

            for index, row in self.excel_data.iterrows():
                if row[0] == self.var_name:
                    count += 1
                    var_skip = index + 1
                    if self.var_name == "CASTELO":
                        var_skip = index + 2
                if row[0] in ["CASTELO", "CID. NOVA", "PLANALTO", "CONTAGEM", "NOVA LIMA", "E-COMM"] and row[0] != self.var_name and count != 0:
                    ind = index - var_skip
                    break
                if self.var_name == "E-COMM" and row[0] != self.var_name and count != 0:
                    last_row = self.excel_data.index[-1]
                    ind = last_row - var_skip + 1
                    break

            if verbose:
                print("var_name =", self.var_name)
                print(row[0])
                print("skip =", var_skip)
                print("ind =", ind)
                print("index =", index)

            self.excel_data = pd.read_excel(file_path, header=None, skiprows=var_skip, nrows=ind)

            if verbose:
                print(f"Valor da coluna {compare_col} na planilha Excel:")
                print(self.excel_data.iloc[:, compare_col])

            return self.excel_data
        
        except FileNotFoundError:
            print(f"File not found: {file_path}")
            messagebox.showerror("Erro", "File not found: {file_path}")
            return None
        except pd.errors.ParserError as e:
            print(f"Parser error occurred while reading the Excel spreadsheet: {e}")
            messagebox.showerror("Erro", "Parser error occurred while reading the Excel")
            return None
        except Exception as e:
            print(f"An error occurred while reading the Excel spreadsheet: {e}")
            messagebox.showerror("Erro", "An error occurred while reading the Excel")
            return None
    
def csv_read(self, file_path, verbose=False):
    compare_col = 'Total'
    print(file_path)
    try:
        self.csv_data = pd.read_csv(file_path, encoding='utf-8', delimiter=';', header=None, skip_blank_lines=False, names=['Total'])

        self.csv_data = self.csv_data.dropna(subset=['Total']).apply(lambda x: x.str.strip() if x.dtype == 'object' else x)

        self.csv_data.to_excel('w3_01-12.xlsx', index=False)
        self.csv_data.to_csv('teste3.csv', index= False)

        self.csv_data = pd.read_excel('w3_01-12.xlsx', header=None, names=['Total'], skiprows=2)
        self.csv_data[compare_col] = pd.to_numeric(self.csv_data[compare_col].replace({r'\.': '', r',': '.'}, regex=True), errors='coerce').fillna(1)

        if verbose:
            print(f"Valores da coluna {compare_col} na planilha convertida:")
            print(self.csv_data[compare_col])

        return self.csv_data
    
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        messagebox.showerror("Erro", "File not found: {file_path}")
        return None
    except pd.errors.ParserError as e:
        print(f"Parser error occurred while reading the CSV spreadsheet: {e}")
        messagebox.showerror("Erro", "Parser error occurred while reading the CSV")
        return None
    except Exception as e:
        print(f"An error occurred while reading the CSV spreadsheet: {e}")
        messagebox.showerror("Erro", "An error occurred while reading the CSV")
        return None
    
def general_read(self, file_path):
    self.general_data = pd.read_excel(file_path, header=None, skiprows= 1)


def adjust_size(self):
    num_rows = max(self.excel_data.shape[0], self.csv_data.shape[0])

    missing_rows = num_rows - self.excel_data.shape[0]
    if missing_rows > 0:
        null_rows = pd.DataFrame(index=range(self.excel_data.shape[0], num_rows))
        self.excel_data = pd.concat([self.excel_data, null_rows])
    return self.excel_data


def process(self, col_w3, col_rede, general, verbose=True,): 
        self.excel_data = adjust_size(self)
        if general:
            general_read(self, 'excel_diferencas_form.xlsx')
            Valor_REDE = self.general_data[col_rede]
            Valor_w3rp = self.general_data[col_w3]
        else:
            Valor_REDE = self.excel_data[col_rede]
            Valor_w3rp = self.csv_data[col_w3]
        
             

        if verbose:
            print("Valores Rede:\n",Valor_REDE)
            print("Valores W3erp:\n",Valor_w3rp)
        
        check_diff(self, Valor_w3rp, Valor_REDE, "w3erp")
        check_diff(self, Valor_REDE, Valor_w3rp, "REDE")
        Checando_pares(self, Valor_REDE, Valor_w3rp)

        difference_sheet = pd.DataFrame(columns=["Data Recebimento", "Data Original", "Valor_REDE", "Valor_w3rp", "Metodo de Pagamento", "Parcelas", "Diferenca"])

        num_linhas = self.csv_data.shape[0]
        for index in range(num_linhas):
            Valor_REDE_atual = Valor_REDE.iloc[index]
            Valor_w3rp_atual = Valor_w3rp.iloc[index]

            diferenca = abs(Valor_REDE_atual - Valor_w3rp_atual)

            nova_linha = {
                "Data Recebimento": self.excel_data.iloc[index, 0],
                "Data Original": self.excel_data.iloc[index, 1],
                "Valor_REDE": Valor_REDE_atual,
                "Valor_w3rp": Valor_w3rp_atual,
                "Metodo de Pagamento": self.excel_data.iloc[index, 9],
                "Parcelas":  self.excel_data.iloc[index, 11],  
                "Diferenca": diferenca
            }

            difference_sheet = difference_sheet._append(nova_linha, ignore_index=True)

        formatar_planilha_diferencas(self, difference_sheet, 'excel_diferencas_form.xlsx')

        if verbose:
            print("Tabela de Diferenças:")
            print(difference_sheet)

        difference_sheet.to_excel('excel_diferencas.xlsx', index=False, float_format="%.2f")

def check_diff(self, coluna_repetidos, coluna_checagem, storage, verbose=False):
    valores = []
    somas_repetidas = {}

    tolerancia = 0.03

    for elemento in coluna_repetidos:
        elemento_arredondado = round(elemento, 2)
        elemento_str = str(elemento_arredondado)

        encontrado = False
        for chave, valor in somas_repetidas.items():
            chave_arredondada = round(float(chave), 2)
            if abs(elemento_arredondado - chave_arredondada) <= tolerancia:
                somas_repetidas[chave] += elemento
                valores.append(elemento)
                encontrado = True
                break

        if not encontrado:
            somas_repetidas[elemento_str] = elemento

    for valor, soma in somas_repetidas.items():
        repeticoes = coluna_repetidos[abs(coluna_repetidos.astype(float) - float(valor)) <=0.03].count()

        if repeticoes >= 2:
            print(f"Repetição do {valor}: {repeticoes} vezes e gera a soma: {soma}")

            indices_checagem = coluna_checagem[abs(coluna_checagem - soma) <= 0.3].index #+ var_skip
            
            if not indices_checagem.empty:
                print(f"A soma {soma} está nas linhas {indices_checagem.to_list()} da REDE")
                if storage == "w3erp":
                    self.w3_storage.append(float(valor))
                    self.rede_storage_s.append(indices_checagem.to_list()[0])
                elif storage == "REDE":
                    self.rede_storage.append(float(valor))
                    self.w3_storage_s.append(indices_checagem.to_list()[0])
            else:
                print(f"A soma {soma} não foi encontrada na coluna_checagem")

    if verbose:
        print("Rede storage: ", self.rede_storage)
        print("W3 storage: ", self.w3_storage)
        print("Rede soma: ", self.rede_storage_s)
        print("W3 soma:", self.w3_storage_s)
        print("Soma repetidos", self.somas_repetidas)

def Checando_pares(self, coluna_REDE, coluna_w3rp, verbose=False):
    sem_par_REDE = []
    sem_par_w3rp = []

    for i_REDE, valor_REDE in enumerate(coluna_REDE):
        encontrado = False
        start_index_w3rp = self.pares_encontrados[-1][1] + 1 if self.pares_encontrados else 0

        for i_w3rp in range(start_index_w3rp, len(coluna_w3rp)):
            valor_w3rp = coluna_w3rp[i_w3rp]

            if abs(valor_REDE - valor_w3rp) < 0.05:
                self.pares_encontrados.append((i_REDE, i_w3rp))
                encontrado = True
                break

        if not encontrado:
            sem_par_REDE.append((i_REDE, valor_REDE))
            sem_par_w3rp.append(None)
    if verbose:
        print("Pares Encontrados:")
        for par in self.pares_encontrados:
            i_REDE, i_w3rp = par
            valor_REDE = coluna_REDE[i_REDE]
            valor_w3rp = coluna_w3rp[i_w3rp]
            print(f"({i_REDE}: {valor_REDE}, {i_w3rp}: {valor_w3rp})")

        print("\nSem Par na coluna_REDE:")
        for sem_par in sem_par_REDE:
            i_REDE, valor_REDE = sem_par
            print(f"({i_REDE}: {valor_REDE}, None)")

        print("\nSem Par na coluna_w3rp:")
        for sem_par in sem_par_w3rp:
            print(f"None")

        print("####################")
        print(self.pares_encontrados)
        return self.pares_encontrados, sem_par_REDE, sem_par_w3rp



def formatar_planilha_diferencas(self, diferencas, caminho_arquivo):
        wb = Workbook()
        ws = wb.active

        cabeçalhos = list(diferencas.columns)
        ws.append(cabeçalhos)

        for r_idx, row in enumerate(diferencas.values, start=2):  
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
            ws[f'G{r_idx}'] = f'=ABS(C{r_idx}-D{r_idx})'

        for col_idx, col_name in enumerate(diferencas.columns, start=1):
            max_length = max(len(str(col_name)), max(len(str(cell.value)) for cell in ws[col_idx]))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        for cell in ws['G']:
            cell.font = Font(bold=True)

        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        ws.conditional_formatting.add('G2:G{0}'.format(len(diferencas) + 1),
                                    CellIsRule(operator='greaterThan', formula=['0.6'], stopIfTrue=True, fill=red_fill))
        
        for linha_numero, linha in enumerate(ws.iter_rows(min_row=2, min_col=3, max_row=ws.max_row, max_col=3), start=2):
            for celula in linha:
                valor = celula.value

                if valor in self.rede_storage:
                    celula.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

                if linha_numero in self.rede_storage_s:
                    celula.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for linha_numero, linha in enumerate(ws.iter_rows(min_row=2, min_col=4, max_row=ws.max_row, max_col=4), start=2):
            for celula in linha:
                valor = celula.value

                if valor in self.w3_storage:
                    celula.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                if linha_numero in self.w3_storage_s:
                    celula.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

        cor_cinza_claro = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cor_cinza_escuro = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")

        for linha_numero, linha in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            if any((linha_numero-2) == par[0] for par in self.pares_encontrados):
                cor = cor_cinza_claro if linha_numero % 2 == 0 else cor_cinza_escuro
                ws.cell(row=linha_numero, column=3).fill = cor

            if any((linha_numero-2) == par[1] for par in self.pares_encontrados):
                cor = cor_cinza_claro if linha_numero % 2 == 0 else cor_cinza_escuro
                ws.cell(row=linha_numero, column=4).fill = cor
                
        wb.save(caminho_arquivo)

def open_sheets(self):
        arquivo_diferencas = 'excel_diferencas_form.xlsx'
        try:
            subprocess.Popen(['start', 'excel', arquivo_diferencas], shell=True)
        except Exception as e:
            print(f"Erro ao abrir o arquivo: {e}")