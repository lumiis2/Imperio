import tkinter
import tkinter.messagebox
import customtkinter
import pandas as pd
import os
import subprocess   
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import codecs
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule


customtkinter.set_appearance_mode("Dark")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("dark-blue")  # Themes: "blue" (standard), "green", "dark-blue"


class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        # configure window
        self.title("Comparador de Planilhas")
        self.geometry(f"{550}x{290}")

        self.grid_columnconfigure(0, weight=1)  
        self.grid_rowconfigure(0, weight=0)  
        self.grid_rowconfigure((1, 2, 3), weight=1)

        self.menubar_frame = customtkinter.CTkFrame(self, height=30, corner_radius=0)
        self.menubar_frame.grid(row=0, column=0, columnspan=3, sticky="nsew")
        self.menubar_frame.grid_rowconfigure(0, weight=1)

        self.chooser_label = customtkinter.CTkLabel(self.menubar_frame, text="Escolha uma loja:", anchor="w")
        self.chooser_label.grid(row=0, column=1, columnspan=2, padx=5, pady=(10, 10))
        self.chooser_optionemenu = customtkinter.CTkOptionMenu(self.menubar_frame, values=["Loja Castelo", "Loja Cidade Nova", "Loja Planalto", "Loja Contagem", "Loja Nova Lima", "Loja E-commerce"],
                                                                       command=lambda selected_store: self.choose_store(selected_store))
        self.chooser_optionemenu.grid(row=0, column=3, columnspan=4, padx=5, pady=(10, 10))

        self.button_1 = customtkinter.CTkButton(self, text="Adicionar REDE", command=lambda: self.download('xlsx'))
        self.button_1.grid(row=1, column=0, padx=20, pady=10)
        self.button_2 = customtkinter.CTkButton(self, text="Adicionar w3erp", command=lambda: self.download('csv'))
        self.button_2.grid(row=2, column=0, padx=20, pady=10)
        


    def choose_store(self, store):
        global var_name
        if store == "Loja Castelo":
            var_name = "CASTELO"
        if store == "Loja Cidade Nova":
            var_name = "CID. NOVA"
        if store == "Loja Planalto":
            var_name = "PLANALTO"
        if store == "Loja Contagem":
            var_name = "CONTAGEM"
        if store =="Loja Nova Lima":
            var_name = "NOVA LIMA"
        if store =="Loja E-commerce":
            var_name = "E-COMM"

    def download(self, type):
        global var_csv, var_xlsx

        if type == 'csv':
            var_csv = filedialog.askopenfilename(title=f"Selecione o arquivo {type.upper()}", filetypes=[(f"Arquivos {type.upper()}", f"*.{type}")])
        elif type == 'xlsx':
            var_xlsx = filedialog.askopenfilename(title=f"Selecione o arquivo {type.upper()}", filetypes=[(f"Arquivos {type.upper()}", f"*.{type}")])

        if var_csv and var_xlsx:
            self.button_3 = customtkinter.CTkButton(self, text="Comparar", command=lambda: self.comparer())
            self.button_3.grid(row=3, column=0, padx=20, pady=10)

    def comparer(self):
        self.process()
        self.open_sheets()

    def process(self):
        global var_csv, var_xlsx

        excel_data = self.excel_read(var_xlsx)
        csv_data = self.csv_read(var_csv)   

        Valor_REDE = excel_data[2]
        Valor_w3rp = csv_data['Total']     
        print(Valor_REDE)
        print(Valor_w3rp)
        
        #check_diff(Valor_w3rp, Valor_REDE)
        print("----------------------------------")
        check_diff(Valor_REDE, Valor_w3rp, "REDE")
        
        #print("REDE para W3")
        print("----------------------------------")
        check_diff(Valor_w3rp, Valor_REDE, "w3erp")
        #Checando_pares(Valor_REDE, Valor_w3rp)

        difference_sheet = pd.DataFrame(columns=["Data Recebimento", "Data Original", "Valor_REDE", "Valor_w3rp", "Metodo de Pagamento", "Parcelas", "Diferenca"])

        # Use min() para garantir que você está iterando sobre o menor número de linhas entre os dois DataFrames
        num_linhas = min(excel_data.shape[0], csv_data.shape[0])

        for index in range(num_linhas):
            Valor_REDE_atual = excel_data.iloc[index, 2]
            Valor_w3rp_atual = csv_data['Total'].iloc[index]

            diferenca = abs(Valor_REDE_atual - Valor_w3rp_atual)

            nova_linha = {
                "Data Recebimento": excel_data.iloc[index, 0],
                "Data Original": excel_data.iloc[index, 1],
                "Valor_REDE": Valor_REDE_atual,
                "Valor_w3rp": Valor_w3rp_atual,
                "Metodo de Pagamento": excel_data.iloc[index, 9],
                "Parcelas":  excel_data.iloc[index, 11],  # Certifique-se de entender como você deseja preencher esta coluna
                "Diferenca": diferenca
            }

            difference_sheet = difference_sheet._append(nova_linha, ignore_index=True)

        self.formatar_planilha_diferencas(difference_sheet, 'Planilhas/excel_diferencas_form.xlsx')

        # Imprimir a tabela de diferenças
        print("Tabela de Diferenças:")
        print(difference_sheet)

        difference_sheet.to_excel('Planilhas/excel_diferencas.xlsx', index=False, float_format="%.2f")
    
    

    def formatar_planilha_diferencas(self, diferencas, caminho_arquivo):
        global repetidos
        global rede_storage
        global w3_storage
        global w3_storage_s
        global rede_storage_s
        wb = Workbook()
        ws = wb.active

        cabeçalhos = list(diferencas.columns)
        ws.append(cabeçalhos)

        # Adicionar os dados ao Excel
        for r_idx, row in enumerate(diferencas.values, start=2):  # Iniciar a partir da segunda linha (após o cabeçalho)
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
            ws[f'G{r_idx}'] = f'=ABS(C{r_idx}-D{r_idx})'

        # Ajustar a largura das colunas com base no tamanho do texto nos cabeçalhos
        for col_idx, col_name in enumerate(diferencas.columns, start=1):
            max_length = max(len(str(col_name)), max(len(str(cell.value)) for cell in ws[col_idx]))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # Definir a coluna "Diferenca" em negrito
        for cell in ws['G']:
            cell.font = Font(bold=True)

        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        ws.conditional_formatting.add('G2:G{0}'.format(len(diferencas) + 1),
                                    CellIsRule(operator='greaterThan', formula=['0.6'], stopIfTrue=True, fill=red_fill))
        
        for linha_numero, linha in enumerate(ws.iter_rows(min_row=2, min_col=3, max_row=ws.max_row, max_col=3), start=2):
            for celula in linha:
                valor = celula.value

                # Destacar em amarelo se o valor está em w3_storage
                if valor in rede_storage:
                    celula.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

                # Destacar em azul se o número da linha está em w3_storage_s
                if linha_numero in rede_storage_s:
                    celula.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for linha_numero, linha in enumerate(ws.iter_rows(min_row=2, min_col=4, max_row=ws.max_row, max_col=4), start=2):
            for celula in linha:
                valor = celula.value

                # Destacar em amarelo se o valor está em w3_storage
                if valor in w3_storage:
                    celula.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                # Destacar em azul se o número da linha está em w3_storage_s
                if linha_numero in w3_storage_s:
                    celula.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        
        # Salvar o arquivo Excel
        wb.save(caminho_arquivo)

    def open_sheets(self):
        arquivo_diferencas = 'Planilhas/excel_diferencas_form.xlsx'

        try:
            # Abre o arquivo com o aplicativo padrão
            subprocess.Popen(['start', 'excel', arquivo_diferencas], shell=True)
        except Exception as e:
            print(f"Erro ao abrir o arquivo: {e}")


    def excel_read(self, file_path):
        global var_name
        compare_col = 2
        global var_skip
        try:

            excel_data = pd.read_excel(file_path, header=None)

            for index, row in excel_data.iterrows():
                if row[0] == var_name:
                    var_skip = index + 1
                    if var_name == "CASTELO":
                        var_skip = index + 2
                if row[0] in ["CASTELO", "CID. NOVA", "PLANALTO", "CONTAGEM", "NOVA LIMA", "E-COMM"] and row[0] != var_name:
                    i = index - var_skip
                    break

            print(var_skip)
            print("i =", i)
            excel_data = pd.read_excel(file_path, header=None, skiprows= var_skip, nrows= i)
            
            print(f"Valor da coluna {compare_col} na planilha Excel:")
            print(excel_data.iloc[:, compare_col])

            return excel_data

        except Exception as e:
            print(f"Ocorreu um erro ao ler a planilha Excel: {e}")
            return None
    

    def change_to_utf8(self, input_path, output_path):
        #NAO DEU CERTO - TESTAR MAIS DEPOIS
        try:
            # Abre o arquivo CSV original para leitura
            with open(input_path, 'r', encoding='ISO-8859-1') as file:
                # Lê as linhas do arquivo CSV
                lines = file.readlines()

            # Abre o novo arquivo CSV UTF-8 para escrita
            with codecs.open('Planilhas/output.txt', 'w', encoding='utf-8') as new_file:
                # Escreve as linhas no novo arquivo CSV UTF-8
                new_file.writelines(lines)

            account = pd.read_csv("Planilhas/output.txt", 
                                delimiter = ';') 
            account.to_csv(output_path, 
                        index = None, encoding='utf-8')

            print(f"Arquivo CSV aberto e salvo com sucesso em: {output_path}")
            return output_path

        except Exception as e:
            print(f"Ocorreu um erro: {e}")

    def csv_read(self, file_path):
        compare_col = 'Total'
        #file_path = self.change_to_utf8(file_path, 'Planilhas/output.csv')
        print(file_path)
        try:
            csv_data = pd.read_csv(file_path, encoding='utf-8', delimiter=';', header=None, skip_blank_lines=False, names=['Total'])

            csv_data = csv_data.dropna(subset=['Total']).apply(lambda x: x.str.strip() if x.dtype == 'object' else x)

            csv_data.to_excel('Planilhas/w3_01-12.xlsx', index=False)
            csv_data.to_csv('Planilhas/teste3.csv', index= False)

            csv_data = pd.read_excel('Planilhas/w3_01-12.xlsx', header=None, names=['Total'], skiprows=2)

            print(f"Valores da coluna {compare_col} na planilha convertida:")
            csv_data[compare_col] = pd.to_numeric(csv_data[compare_col].replace({r'\.': '', r',': '.'}, regex=True), errors='coerce').fillna(1)


            print(csv_data[compare_col])

            return csv_data
        
        except Exception as e:
            print(f"Ocorreu um erro ao ler a planilha CSV: {e}")
            return None
        


def check_diff(coluna_repetidos, coluna_checagem, storage):
    global var_skip
    # Criar um dicionário para armazenar as somas dos valores repetidos
    global rede_storage
    global w3_storage
    global w3_storage_s
    global rede_storage_s
    somas_repetidas = {}

    # Iterar sobre os elementos da coluna
    for elemento in coluna_repetidos:
        elemento_str = str(round(elemento, 2))

        # Adicionar o valor ao dicionário ou somar se já existir
        if elemento_str in somas_repetidas:
            somas_repetidas[elemento_str] += elemento
        else:
            somas_repetidas[elemento_str] = elemento

    # Iterar sobre o dicionário e imprimir as repetições e somas
    for valor, soma in somas_repetidas.items():
        repeticoes = coluna_repetidos[coluna_repetidos.astype(str) == valor].count()

        if repeticoes >= 2:
            print(f"Repetição do {valor}: {repeticoes} vezes e gera a soma: {soma}")

            indices_checagem = coluna_checagem[abs(coluna_checagem - soma) <= 0.3].index + var_skip
            
            # Imprimir os índices correspondentes
            if not indices_checagem.empty:
                print(f"A soma {soma} está nas linhas {indices_checagem.to_list()} da REDE")
                if storage == "w3erp":
                    w3_storage.append(float(valor))
                    rede_storage_s.append(indices_checagem.to_list()[0])
                elif storage == "REDE":
                    rede_storage.append(float(valor))
                    w3_storage_s.append(indices_checagem.to_list()[0])
            else:
                print(f"A soma {soma} não foi encontrada na coluna_checagem")

    print("rede storage", rede_storage)
    print(w3_storage)
    print(rede_storage_s)
    print(w3_storage_s)

if __name__ == "__main__":

    script_directory = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_directory)

    var_csv = ''
    var_xlsx = ''
    root = None 
    var_name = "CASTELO"
    var_skip = 0
    var_tolerance = 0.5
    rede_storage = []
    rede_storage_s = []
    w3_storage = []
    w3_storage_s = []
    app = App()
    app.mainloop()