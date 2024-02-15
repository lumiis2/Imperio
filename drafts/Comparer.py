import tkinter
import customtkinter
import subprocess   
import codecs
import os
import tkinter.messagebox
import pandas as pd
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule


customtkinter.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("dark-blue")  # Themes: "blue" (standard), "green", "dark-blue"


class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        self.title("Comparador de Planilhas")
        self.geometry(f"{550}x{290}")

        self.grid_columnconfigure(0, weight=1)  
        self.grid_rowconfigure(0, weight=0)  
        self.grid_rowconfigure((1, 2, 3), weight=1)

        self.menubar_frame = customtkinter.CTkFrame(self, height=30, corner_radius=0)
        self.menubar_frame.grid(row=0, column=0, columnspan=3, sticky="nsew")
        self.menubar_frame.grid_rowconfigure(0, weight=1)

        self.chooser_label = customtkinter.CTkLabel(self.menubar_frame, text="Escolha uma loja:", anchor="w")
        self.chooser_label.grid(row=0, column=1, columnspan=2, padx=5, pady=(10, 10))
        self.chooser_optionemenu = customtkinter.CTkOptionMenu(self.menubar_frame, values=["Loja Castelo", "Loja Cidade Nova", "Loja Planalto", "Loja Contagem", "Loja Nova Lima", "Loja E-commerce"],
                                                                       command=lambda selected_store: self.choose_store(selected_store))
        self.chooser_optionemenu.grid(row=0, column=3, columnspan=4, padx=5, pady=(10, 10))

        self.button_1 = customtkinter.CTkButton(self, text="Adicionar REDE", command=lambda: self.download('xlsx'))
        self.button_1.grid(row=1, column=0, padx=20, pady=10)
        self.button_2 = customtkinter.CTkButton(self, text="Adicionar w3erp", command=lambda: self.download('csv'))
        self.button_2.grid(row=2, column=0, padx=20, pady=10)

        self.switch = customtkinter.CTkSwitch(master=self.menubar_frame, text="DarkMode", command=lambda: self.change_mode())
        self.switch.grid(row=0, column=10, padx=30, pady=(5, 5))
        
    def change_mode(self):
        global mode
        mode += 1
        customtkinter.set_appearance_mode("Light") if mode % 2 == 0 else customtkinter.set_appearance_mode("Dark")


    def choose_store(self, store):
        global var_name
        if store == "Loja Castelo":
            var_name = "CASTELO"
        if store == "Loja Cidade Nova":
            var_name = "CID. NOVA"
        if store == "Loja Planalto":
            var_name = "PLANALTO"
        if store == "Loja Contagem":
            var_name = "CONTAGEM"
        if store =="Loja Nova Lima":
            var_name = "NOVA LIMA"
        if store =="Loja E-commerce":
            var_name = "E-COMM"

    def download(self, type):
        global var_csv, var_xlsx

        if type == 'csv':
            var_csv = filedialog.askopenfilename(title=f"Selecione o arquivo {type.upper()}", filetypes=[(f"Arquivos {type.upper()}", f"*.{type}")])
        elif type == 'xlsx':
            var_xlsx = filedialog.askopenfilename(title=f"Selecione o arquivo {type.upper()}", filetypes=[(f"Arquivos {type.upper()}", f"*.{type}")])

        if var_csv and var_xlsx:
            self.button_3 = customtkinter.CTkButton(self, text="Comparar", command=lambda: self.comparer())
            self.button_3.grid(row=3, column=0, padx=20, pady=10)

    def comparer(self):
        self.process()
        self.open_sheets()
        self.button_3 = customtkinter.CTkButton(self, text="Rodar Novamente", command=lambda: self.rodar())
        self.button_3.grid(row=4, column=0, padx=20, pady=10)

    def rodar(self, verbose=False):
        data = pd.read_excel('excel_diferencas_form.xlsx', header=None, skiprows=1)
        Valor_REDE = data[2]
        Valor_w3rp = data[3]
        print(Valor_REDE)
        print(Valor_w3rp)
        check_diff(Valor_w3rp, Valor_REDE, "w3erp")
        check_diff(Valor_REDE, Valor_w3rp, "REDE")
        Checando_pares(Valor_REDE, Valor_w3rp)
        self.formatar_planilha_diferencas(data, 'excel_diferencas_form.xlsx')

    def ajustar_tamanho(self, excel_data, csv_data):
        # Obtém o número máximo de linhas entre os dois DataFrames
        num_linhas = max(excel_data.shape[0], csv_data.shape[0])

        # Adiciona linhas nulas no DataFrame do Excel, se necessário
        linhas_faltantes = num_linhas - excel_data.shape[0]
        if linhas_faltantes > 0:
            linhas_nulas = pd.DataFrame(index=range(excel_data.shape[0], num_linhas))
            excel_data = pd.concat([excel_data, linhas_nulas])

        # Agora, os dois DataFrames têm o mesmo número de linhas
        return excel_data


    def process(self, verbose=True):
        global var_csv, var_xlsx
        
        excel_data = self.excel_read(var_xlsx)
        csv_data = self.csv_read(var_csv)   
        excel_data = self.ajustar_tamanho(excel_data, csv_data)
        Valor_REDE = excel_data[2]
        Valor_w3rp = csv_data['Total'] 

        if verbose:
            print("Valores Rede:\n",Valor_REDE)
            print("Valores W3erp:\n",Valor_w3rp)
        
        check_diff(Valor_w3rp, Valor_REDE, "w3erp")
        check_diff(Valor_REDE, Valor_w3rp, "REDE")
        Checando_pares(Valor_REDE, Valor_w3rp)

        difference_sheet = pd.DataFrame(columns=["Data Recebimento", "Data Original", "Valor_REDE", "Valor_w3rp", "Metodo de Pagamento", "Parcelas", "Diferenca"])

        num_linhas = csv_data.shape[0]
        for index in range(num_linhas):
            Valor_REDE_atual = Valor_REDE.iloc[index]
            Valor_w3rp_atual = Valor_w3rp.iloc[index]

            diferenca = abs(Valor_REDE_atual - Valor_w3rp_atual)

            nova_linha = {
                "Data Recebimento": excel_data.iloc[index, 0],
                "Data Original": excel_data.iloc[index, 1],
                "Valor_REDE": Valor_REDE_atual,
                "Valor_w3rp": Valor_w3rp_atual,
                "Metodo de Pagamento": excel_data.iloc[index, 9],
                "Parcelas":  excel_data.iloc[index, 11],  
                "Diferenca": diferenca
            }

            difference_sheet = difference_sheet._append(nova_linha, ignore_index=True)

        self.formatar_planilha_diferencas(difference_sheet, 'excel_diferencas_form.xlsx')

        if verbose:
            print("Tabela de Diferenças:")
            print(difference_sheet)

        difference_sheet.to_excel('excel_diferencas.xlsx', index=False, float_format="%.2f")
    
    

    def formatar_planilha_diferencas(self, diferencas, caminho_arquivo):
        global repetidos
        global rede_storage
        global w3_storage
        global w3_storage_s
        global rede_storage_s
        global pares_encontrados
        wb = Workbook()
        ws = wb.active

        cabeçalhos = list(diferencas.columns)
        ws.append(cabeçalhos)

        for r_idx, row in enumerate(diferencas.values, start=2):  
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
            ws[f'G{r_idx}'] = f'=ABS(C{r_idx}-D{r_idx})'

        for col_idx, col_name in enumerate(diferencas.columns, start=1):
            max_length = max(len(str(col_name)), max(len(str(cell.value)) for cell in ws[col_idx]))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        for cell in ws['G']:
            cell.font = Font(bold=True)

        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        ws.conditional_formatting.add('G2:G{0}'.format(len(diferencas) + 1),
                                    CellIsRule(operator='greaterThan', formula=['0.6'], stopIfTrue=True, fill=red_fill))
        
        for linha_numero, linha in enumerate(ws.iter_rows(min_row=2, min_col=3, max_row=ws.max_row, max_col=3), start=2):
            for celula in linha:
                valor = celula.value

                if valor in rede_storage:
                    celula.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

                if linha_numero in rede_storage_s:
                    celula.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for linha_numero, linha in enumerate(ws.iter_rows(min_row=2, min_col=4, max_row=ws.max_row, max_col=4), start=2):
            for celula in linha:
                valor = celula.value

                if valor in w3_storage:
                    celula.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                if linha_numero in w3_storage_s:
                    celula.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

        cor_cinza_claro = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cor_cinza_escuro = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")

        for linha_numero, linha in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            if any((linha_numero-2) == par[0] for par in pares_encontrados):
                cor = cor_cinza_claro if linha_numero % 2 == 0 else cor_cinza_escuro
                ws.cell(row=linha_numero, column=3).fill = cor

            if any((linha_numero-2) == par[1] for par in pares_encontrados):
                cor = cor_cinza_claro if linha_numero % 2 == 0 else cor_cinza_escuro
                ws.cell(row=linha_numero, column=4).fill = cor
                
        wb.save(caminho_arquivo)

    def open_sheets(self):
        arquivo_diferencas = 'excel_diferencas_form.xlsx'
        try:
            subprocess.Popen(['start', 'excel', arquivo_diferencas], shell=True)
        except Exception as e:
            print(f"Erro ao abrir o arquivo: {e}")


    def excel_read(self, file_path, verbose=True):
        global var_name
        compare_col = 2
        count = 0
        ind = 0
        global var_skip
        try:
            excel_data = pd.read_excel(file_path, header=None)

            for index, row in excel_data.iterrows():
                if row[0] == var_name:
                    print("Entrou 1")
                    count += 1
                    var_skip = index + 1
                    if var_name == "CASTELO":
                        print("Entrou 2")
                        var_skip = index + 2
                if row[0] in ["CASTELO", "CID. NOVA", "PLANALTO", "CONTAGEM", "NOVA LIMA", "E-COMM"] and row[0] != var_name and count != 0:
                    print("Entrou 3")
                    ind = index - var_skip
                    break
                if var_name == "E-COMM" and row[0] != var_name and count != 0:
                    print("Entrou 4")
                    last_row = excel_data.index[-1]
                    print(last_row)
                    ind = last_row - var_skip + 1
                    break

            if verbose:
                print("var_name =", var_name)
                print(row[0])
                print("skip =", var_skip)
                print("ind =", ind)
                print("index =", index)
            
            excel_data = pd.read_excel(file_path, header=None, skiprows= var_skip, nrows= ind)
            
            if verbose:
                print(f"Valor da coluna {compare_col} na planilha Excel:")
                print(excel_data.iloc[:, compare_col])

            return excel_data

        except Exception as e:
            print(f"Ocorreu um erro ao ler a planilha Excel: {e}")
            return None
    

    def change_to_utf8(self, input_path, output_path):
        #NAO DEU CERTO - TESTAR MAIS DEPOIS
        #A LEITURA E A CRIACAO DO TXT ESTA FUNCIONANDO
        try:
            with open(input_path, 'r', encoding='ISO-8859-1') as file:
                lines = file.readlines()

            with codecs.open('output.txt', 'w', encoding='utf-8') as new_file:
                new_file.writelines(lines)

            account = pd.read_csv("output.txt", 
                                delimiter = ';') 
            account.to_csv(output_path, 
                        index = None, encoding='utf-8')
            print(f"Arquivo CSV aberto e salvo com sucesso em: {output_path}")
            return output_path
        
        except Exception as e:
            print(f"Ocorreu um erro: {e}")


    def csv_read(self, file_path, verbose=False):
        compare_col = 'Total'
        print(file_path)
        try:
            csv_data = pd.read_csv(file_path, encoding='utf-8', delimiter=';', header=None, skip_blank_lines=False, names=['Total'])

            csv_data = csv_data.dropna(subset=['Total']).apply(lambda x: x.str.strip() if x.dtype == 'object' else x)

            csv_data.to_excel('w3_01-12.xlsx', index=False)
            csv_data.to_csv('teste3.csv', index= False)

            csv_data = pd.read_excel('w3_01-12.xlsx', header=None, names=['Total'], skiprows=2)
            csv_data[compare_col] = pd.to_numeric(csv_data[compare_col].replace({r'\.': '', r',': '.'}, regex=True), errors='coerce').fillna(1)

            if verbose:
                print(f"Valores da coluna {compare_col} na planilha convertida:")
                print(csv_data[compare_col])

            return csv_data
        
        except Exception as e:
            print(f"Ocorreu um erro ao ler a planilha CSV: {e}")
            return None
        


def check_diff(coluna_repetidos, coluna_checagem, storage, verbose=False):
    global var_skip
    global rede_storage
    global w3_storage
    global w3_storage_s
    global rede_storage_s
    valores = []
    somas_repetidas = {}
    

    tolerancia = 0.03

    for elemento in coluna_repetidos:
        elemento_arredondado = round(elemento, 2)
        elemento_str = str(elemento_arredondado)

        encontrado = False
        for chave, valor in somas_repetidas.items():
            chave_arredondada = round(float(chave), 2)
            if abs(elemento_arredondado - chave_arredondada) <= tolerancia:
                somas_repetidas[chave] += elemento
                valores.append(elemento)
                encontrado = True
                break

        if not encontrado:
            somas_repetidas[elemento_str] = elemento

    for valor, soma in somas_repetidas.items():
        repeticoes = coluna_repetidos[abs(coluna_repetidos.astype(float) - float(valor)) <=0.03].count()

        if repeticoes >= 2:
            print(f"Repetição do {valor}: {repeticoes} vezes e gera a soma: {soma}")

            indices_checagem = coluna_checagem[abs(coluna_checagem - soma) <= 0.3].index + var_skip
            
            if not indices_checagem.empty:
                print(f"A soma {soma} está nas linhas {indices_checagem.to_list()} da REDE")
                if storage == "w3erp":
                    w3_storage.append(float(valor))
                    rede_storage_s.append(indices_checagem.to_list()[0])
                elif storage == "REDE":
                    rede_storage.append(float(valor))
                    w3_storage_s.append(indices_checagem.to_list()[0])
            else:
                print(f"A soma {soma} não foi encontrada na coluna_checagem")

    if verbose:
        print("Rede storage: ", rede_storage)
        print("W3 storage: ", w3_storage)
        print("Rede soma: ", rede_storage_s)
        print("W3 soma:", w3_storage_s)
        print("Soma repetidos", somas_repetidas)



def Checando_pares(coluna_REDE, coluna_w3rp, verbose=False):
    global pares_encontrados

    sem_par_REDE = []
    sem_par_w3rp = []

    for i_REDE, valor_REDE in enumerate(coluna_REDE):
        encontrado = False
        start_index_w3rp = pares_encontrados[-1][1] + 1 if pares_encontrados else 0

        for i_w3rp in range(start_index_w3rp, len(coluna_w3rp)):
            valor_w3rp = coluna_w3rp[i_w3rp]

            if abs(valor_REDE - valor_w3rp) < 0.05:
                pares_encontrados.append((i_REDE, i_w3rp))
                encontrado = True
                break

        if not encontrado:
            sem_par_REDE.append((i_REDE, valor_REDE))
            sem_par_w3rp.append(None)
    if verbose:
        print("Pares Encontrados:")
        for par in pares_encontrados:
            i_REDE, i_w3rp = par
            valor_REDE = coluna_REDE[i_REDE]
            valor_w3rp = coluna_w3rp[i_w3rp]
            print(f"({i_REDE}: {valor_REDE}, {i_w3rp}: {valor_w3rp})")

        print("\nSem Par na coluna_REDE:")
        for sem_par in sem_par_REDE:
            i_REDE, valor_REDE = sem_par
            print(f"({i_REDE}: {valor_REDE}, None)")

        print("\nSem Par na coluna_w3rp:")
        for sem_par in sem_par_w3rp:
            print(f"None")

        print("####################")
        print(pares_encontrados)
        return pares_encontrados, sem_par_REDE, sem_par_w3rp

if __name__ == "__main__":

    script_directory = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_directory)

    var_csv = ''
    var_xlsx = ''
    root = None 
    var_name = "CASTELO"
    var_skip = 0
    var_tolerance = 0.5
    rede_storage = []
    rede_storage_s = []
    w3_storage = []
    w3_storage_s = []
    pares_encontrados = []
    rodar = 0
    mode = 0
    app = App()
    app.mainloop()