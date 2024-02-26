
import pandas as pd
import customtkinter
from tkinter import filedialog
import pandas as pd
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from aux_functions import open_sheets, adjust_size, check_diff, Checando_pares, reset_var

def download(self, file_type):
    if self.order == 0 and (not self.excel_data.empty and not self.csv_data.empty) and self.temp_var_name != self.var_name:
        messagebox.showwarning("ATENÇÃO!!!", "Lembre-se de selecionar a loja.")
        return 1
    self.temp_var_name = self.var_name
    file_path = filedialog.askopenfilename(
        title=f"Select the {file_type.upper()} file",
        filetypes=[(f"{file_type.upper()} files", f"*.{file_type}")]
    )
    if file_path:
        if file_type == 'csv':
            self.csv_data = csv_read(self, file_path)
            self.csv_path = file_path
        elif file_type == 'xlsx':
            self.excel_data = excel_read(self, file_path)
            self.excel_path = file_path

        create_compare_button(self)

def create_compare_button(self):
    if not self.csv_data.empty and not self.excel_data.empty:
        self.button_3 = customtkinter.CTkButton(self, text="Comparar", command=lambda: comparer(self))
        self.button_3.grid(row=3, column=0, padx=20, pady=10)

def reprocess(self, file_type):
    if file_type == 'csv':
        self.csv_data = pd.DataFrame()
    elif file_type == 'excel':
        self.excel_data = pd.DataFrame()
    else:
        messagebox.showerror("Erro", "Tipo de arquivo inválido. Use 'csv' ou 'excel'.")
        raise ValueError("Tipo de arquivo inválido. Use 'csv' ou 'excel'.")
    download(self, file_type)

def excel_read(self, file_path, verbose=True):
    # Definindo variáveis iniciais
    compare_col = 2
    count = 0
    try:
        # Lendo o arquivo Excel
        self.excel_data = pd.read_excel(file_path, header=None)
        print(self.excel_data)

        # Encontrando var_skip
        var_skip = self.excel_data[self.excel_data[0] == self.var_name].index.min() + 1
        if self.var_name == "CASTELO":
            var_skip += 1

        # Encontrando ind
        relevant_rows = self.excel_data.loc[var_skip:, 0]
        ind = relevant_rows.isin(["CASTELO", "CID. NOVA", "PLANALTO", "CONTAGEM", "NOVA LIMA", "E-COMM"]).idxmax()
        if self.var_name == 'E-COMM':
            last_row = self.excel_data.index[-1]
            ind = last_row 

        # Extraindo os dados relevantes
        self.excel_data = self.excel_data.iloc[var_skip:ind]

        if verbose:
            print("var_name =", self.var_name)
            print("skip =", var_skip)
            print("ind =", ind)

        if verbose:
            print(f"Valor da coluna {compare_col} na planilha Excel:")
            print(self.excel_data.iloc[:, compare_col])

        return self.excel_data
        
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        messagebox.showerror("Erro", "File not found: {file_path}")
        reprocess(self, 'excel')
        return self.excel_data
    except pd.errors.ParserError as e:
        print(f"Parser error occurred while reading the Excel spreadsheet: {e}")
        messagebox.showerror("Erro", "Parser error occurred while reading the Excel")
        reprocess(self, 'excel')
        return self.excel_data
    except Exception as e:
        print(f"An error occurred while reading the Excel spreadsheet: {e}")
        messagebox.showerror("Erro", "An error occurred while reading the Excel")
        reprocess(self, 'excel')
        return self.excel_data

    
def csv_read(self, file_path, verbose=False):
    compare_col = 'Total'
    print(file_path)
    try:
        self.csv_data = pd.read_csv(file_path, encoding='utf-8', delimiter=';', header=None, skip_blank_lines=False, names=['Total'])

        self.csv_data = self.csv_data.dropna(subset=['Total']).apply(lambda x: x.str.strip() if x.dtype == 'object' else x)

        self.csv_data.to_excel('w3_01-12.xlsx', index=False)
        self.csv_data.to_csv('teste3.csv', index= False)

        self.csv_data = pd.read_excel('w3_01-12.xlsx', header=None, names=['Total'], skiprows=2)
        self.csv_data[compare_col] = pd.to_numeric(self.csv_data[compare_col].replace({r'\.': '', r',': '.'}, regex=True), errors='coerce').fillna(1)

        if verbose:
            print(f"Valores da coluna {compare_col} na planilha convertida:")
            print(self.csv_data[compare_col])

        return self.csv_data
    
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        messagebox.showerror("Erro", "File not found: {file_path}")
        reprocess(self, 'csv')
        return self.csv_data
    except pd.errors.ParserError as e:
        print(f"Parser error occurred while reading the CSV spreadsheet: {e}")
        messagebox.showerror("Erro", "Parser error occurred while reading the CSV")
        reprocess(self, 'csv')
        return self.csv_data
    except Exception as e:
        print(f"An error occurred while reading the CSV spreadsheet: {e}")
        messagebox.showerror("Erro", "An error occurred while reading the CSV")
        reprocess(self, 'csv')
        return self.csv_data
    
def general_read(self, file_path):
    self.general_data = pd.read_excel(file_path, header=None, skiprows= 1)

def comparer(self):
        general = False
        if self.temp_excel_path == self.excel_path and self.temp_csv_path == self.csv_path:
            general = True
            process(self, 3, 2, general)
            open_sheets(self)
            reset_var(self)
        else:
            process(self, 'Total', 2, general)
            open_sheets(self)
            reset_var(self)
        self.temp_excel_path = self.excel_path
        self.temp_csv_path = self.csv_path

def process(self, col_w3, col_rede, general, verbose=True,): 
        self.excel_data = adjust_size(self)
        if general:
            general_read(self, 'excel_diferencas_form.xlsx')
            Valor_REDE = self.general_data[col_rede]
            Valor_w3rp = self.general_data[col_w3]
        else:
            Valor_REDE = self.excel_data[col_rede]
            Valor_w3rp = self.csv_data[col_w3]

        if verbose:
            print("Valores Rede:\n",Valor_REDE)
            print("Valores W3erp:\n",Valor_w3rp)
        
        check_diff(self, Valor_w3rp, Valor_REDE, "w3erp")
        check_diff(self, Valor_REDE, Valor_w3rp, "REDE")
        Checando_pares(self, Valor_REDE, Valor_w3rp)

        difference_sheet = pd.DataFrame()
        difference_sheet = pd.DataFrame(columns=["Data Recebimento", "Data Original", "Valor_REDE", "Valor_w3rp", "Metodo de Pagamento", "Parcelas", "Diferenca"])

        num_linhas = self.csv_data.shape[0]
        for index in range(num_linhas):
            Valor_REDE_atual = Valor_REDE.iloc[index]
            Valor_w3rp_atual = Valor_w3rp.iloc[index]

            diferenca = abs(Valor_REDE_atual - Valor_w3rp_atual)

            nova_linha = {
                "Data Recebimento": self.excel_data.iloc[index, 0],
                "Data Original": self.excel_data.iloc[index, 1],
                "Valor_REDE": Valor_REDE_atual,
                "Valor_w3rp": Valor_w3rp_atual,
                "Metodo de Pagamento": self.excel_data.iloc[index, 9],
                "Parcelas":  self.excel_data.iloc[index, 11],  
                "Diferenca": diferenca
            }

            difference_sheet = difference_sheet._append(nova_linha, ignore_index=True)

        formatar_planilha_diferencas(self, difference_sheet, 'excel_diferencas_form.xlsx')

        if verbose:
            print("Tabela de Diferenças:")
            print(difference_sheet)

        difference_sheet.to_excel('excel_diferencas.xlsx', index=False, float_format="%.2f")

def formatar_planilha_diferencas(self, diferencas, caminho_arquivo):
        wb = Workbook()
        ws = wb.active

        cabeçalhos = list(diferencas.columns)
        ws.append(cabeçalhos)

        for r_idx, row in enumerate(diferencas.values, start=2):  
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
            ws[f'G{r_idx}'] = f'=ABS(C{r_idx}-D{r_idx})'

        for col_idx, col_name in enumerate(diferencas.columns, start=1):
            max_length = max(len(str(col_name)), max(len(str(cell.value)) for cell in ws[col_idx]))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        for cell in ws['G']:
            cell.font = Font(bold=True)

        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        ws.conditional_formatting.add('G2:G{0}'.format(len(diferencas) + 1),
                                    CellIsRule(operator='greaterThan', formula=['0.6'], stopIfTrue=True, fill=red_fill))
        
        for linha_numero, linha in enumerate(ws.iter_rows(min_row=2, min_col=3, max_row=ws.max_row, max_col=3), start=2):
            for celula in linha:
                valor = celula.value

                if valor in self.rede_storage:
                    celula.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

                if linha_numero in self.rede_storage_s:
                    celula.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for linha_numero, linha in enumerate(ws.iter_rows(min_row=2, min_col=4, max_row=ws.max_row, max_col=4), start=2):
            for celula in linha:
                valor = celula.value

                if valor in self.w3_storage:
                    celula.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                if linha_numero in self.w3_storage_s:
                    celula.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

        cor_cinza_claro = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cor_cinza_escuro = PatternFill(start_color="A9A9A9", end_color="A9A9A9", fill_type="solid")

        for linha_numero, linha in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            if any((linha_numero-2) == par[0] for par in self.pares_encontrados):
                cor = cor_cinza_claro if linha_numero % 2 == 0 else cor_cinza_escuro
                ws.cell(row=linha_numero, column=3).fill = cor

            if any((linha_numero-2) == par[1] for par in self.pares_encontrados):
                cor = cor_cinza_claro if linha_numero % 2 == 0 else cor_cinza_escuro
                ws.cell(row=linha_numero, column=4).fill = cor
                
        wb.save(caminho_arquivo)
