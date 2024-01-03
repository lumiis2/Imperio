import pandas as pd
import os
import subprocess   
from tkinter import Tk, Button, Label, Toplevel, Menu, filedialog



def criar_janela():
    global root

    root = Tk()
    root.title("Comparador de Tabelas")
    root.geometry("400x400")  # Largura x Altura

     # Cria uma barra de menu
    barra_menu = Menu(root)
    root.config(menu=barra_menu)

    # Adiciona um item "Loja" à barra de menu
    menu_loja = Menu(barra_menu, tearoff=0)
    lojas = ["Loja Castelo", "Loja Cidade Nova", "Loja Planalto", "Loja Contagem", "Loja Nova Lima", "Loja E-commerce"]

    # Cria um submenu para a opção "Loja"
    submenu_loja = Menu(menu_loja, tearoff=0)

    # Adiciona os nomes das lojas como comandos no submenu
    for loja in lojas:
        submenu_loja.add_command(label=loja, command=lambda l=loja: escolher_loja(l, menu_loja))

    barra_menu.add_cascade(label="Loja", menu=submenu_loja)

    btn_xlsx = Button(root, text="Adicionar XLSX", command=lambda: carregar_arquivo('xlsx'), width=20, height=3)
    btn_xlsx.pack(pady=20)

    btn_csv = Button(root, text="Adicionar CSV", command=lambda: carregar_arquivo('csv'), width=20, height=3)
    btn_csv.pack(pady=20)

    root.mainloop()

def carregar_arquivo(tipo):
    global var_csv, var_xlsx

    if tipo == 'csv':
        var_csv = filedialog.askopenfilename(title=f"Selecione o arquivo {tipo.upper()}", filetypes=[(f"Arquivos {tipo.upper()}", f"*.{tipo}")])
    elif tipo == 'xlsx':
        var_xlsx = filedialog.askopenfilename(title=f"Selecione o arquivo {tipo.upper()}", filetypes=[(f"Arquivos {tipo.upper()}", f"*.{tipo}")])

    if var_csv and var_xlsx:
        btn_run = Button(root, text="Run", command=executar_comparacao, width=20, height=3)
        btn_run.pack(pady=20)

def executar_comparacao():
    processar_arquivos()
    abrir_planilha_diferencas()


def escolher_loja(name, menu_loja):
    # Escrever na barra superior de menu ao lado de "Loja"
    menu_loja.entryconfigure(0, label=f"Loja: {name}")

    # Ler cada linha da planilha até encontrar o nome
    # Substitua isso pela lógica real de leitura da planilha
    with open('sua_planilha.csv', 'r') as file:
        for linha in file:
            if name in linha:
                # Imprime a linha no terminal
                print(f"Linha encontrada: {linha}")
                # Retorna o nome quando encontrado
                return name

def ler_planilha_excel(file_path):
    compare_col = 2
    try:
        #o numero de skiprows pode ser o numero da linha do titulo para poder pegar de cada loja corretamente
        #PENSAR EM COMO IMPLEMENTAR ESSA LOGICA
        #PENSAR COMO LIDAR COM BAIXAS CRED E DEBITO PRA JA TIRAR
        excel_data = pd.read_excel(file_path, header=None, skiprows=221)
        
        print(f"Valor da coluna {compare_col} na planilha Excel:")
        print(excel_data.iloc[:, compare_col])

        return excel_data

    except Exception as e:
        print(f"Ocorreu um erro ao ler a planilha Excel: {e}")
        return None

def ler_planilha_csv(file_path):
    compare_col = 'Total'
    try:
        csv_data = pd.read_csv(file_path, encoding='utf-8', delimiter=';', header=None, skip_blank_lines=False, names=['Total'])

        csv_data = csv_data.dropna(subset=['Total']).apply(lambda x: x.str.strip() if x.dtype == 'object' else x)

        csv_data.to_excel('Planilhas/w3_01-12.xlsx', index=False)

        csv_data = pd.read_excel('Planilhas/w3_01-12.xlsx', header=None, names=['Total'], skiprows=2)

        print(f"Valores da coluna {compare_col} na planilha convertida:")
        csv_data[compare_col] = pd.to_numeric(csv_data[compare_col].replace({r'\.': '', r',': '.'}, regex=True), errors='coerce').fillna(1)


        print(csv_data[compare_col])

        return csv_data

    except Exception as e:
        print(f"Ocorreu um erro ao ler a planilha CSV: {e}")
        return None


def processar_arquivos():
    global var_csv, var_xlsx
    print(var_xlsx)
    print(var_csv)

    excel_data = ler_planilha_excel(var_xlsx)

    csv_data = ler_planilha_csv(var_csv)

    Valor_REDE = excel_data[2]
    Valor_w3rp = csv_data['Total']
    Parcelas = excel_data[11]\
    

    print("W3 para REDE")
    check_diff(Valor_w3rp, Valor_REDE)
    #print("REDE para W3")
    #check_diff(Valor_REDE, Valor_w3rp)
    #Checando_pares(Valor_REDE, Valor_w3rp)

    difference_sheet = pd.DataFrame(columns=["Data Recebimento", "Data Original", "Valor_REDE", "Valor_w3rp", "Metodo de Pagamento", "Parcelas", "Diferenca"])

    # Use min() para garantir que você está iterando sobre o menor número de linhas entre os dois DataFrames
    num_linhas = min(excel_data.shape[0], csv_data.shape[0])

    for index in range(num_linhas):
        Valor_REDE_atual = excel_data.iloc[index, 2]
        Valor_w3rp_atual = Valor_w3rp.iloc[index]

        diferenca = abs(Valor_REDE_atual - Valor_w3rp_atual)

        nova_linha = {
            "Data Recebimento": excel_data.iloc[index, 0],
            "Data Original": excel_data.iloc[index, 1],
            "Valor_REDE": Valor_REDE_atual,
            "Valor_w3rp": Valor_w3rp_atual,
            "Metodo de Pagamento": excel_data.iloc[index, 9],
            "Parcelas":  excel_data.iloc[index, 11],  # Certifique-se de entender como você deseja preencher esta coluna
            "Diferenca": diferenca
        }

        difference_sheet = difference_sheet._append(nova_linha, ignore_index=True)

    formatar_planilha_diferencas(difference_sheet, 'Planilhas/excel_diferencas_form.xlsx')

    # Imprimir a tabela de diferenças
    print("Tabela de Diferenças:")
    print(difference_sheet)

    # Salvar a tabela de diferenças como XLSX
    difference_sheet.to_excel('Planilhas/excel_diferencas.xlsx', index=False, float_format="%.2f")

    
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

def formatar_planilha_diferencas(diferencas, caminho_arquivo):
    # Criar um novo arquivo Excel
    wb = Workbook()
    ws = wb.active

    # Adicionar os nomes das colunas como a primeira linha (cabeçalho)
    cabeçalhos = list(diferencas.columns)
    ws.append(cabeçalhos)

    # Adicionar os dados ao Excel
    for r_idx, row in enumerate(diferencas.values, start=2):  # Iniciar a partir da segunda linha (após o cabeçalho)
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

    # Ajustar a largura das colunas com base no tamanho do texto nos cabeçalhos
    for col_idx, col_name in enumerate(diferencas.columns, start=1):
        max_length = max(len(str(col_name)), max(len(str(cell.value)) for cell in ws[col_idx]))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Definir a coluna "Diferenca" em negrito
    for cell in ws['G']:
        cell.font = Font(bold=True)

    # Destacar valores de diferença acima de 60 centavos em vermelho
    for row in ws.iter_rows(min_row=2, max_col=7, max_row=ws.max_row):
        for cell in row:
            if cell.column == 7:  # Coluna "Diferenca"
                if isinstance(cell.value, (int, float)) and cell.value > 0.6:  # 60 centavos
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Salvar o arquivo Excel
    wb.save(caminho_arquivo)
def abrir_planilha_diferencas():
    arquivo_diferencas = 'Planilhas/excel_diferencas_form.xlsx'

    try:
        # Abre o arquivo com o aplicativo padrão
        subprocess.Popen(['start', 'excel', arquivo_diferencas], shell=True)
    except Exception as e:
        print(f"Erro ao abrir o arquivo: {e}")
    

def check_diff(coluna_repetidos, coluna_checagem):
    # Criar um dicionário para armazenar as somas dos valores repetidos
    somas_repetidas = {}

    # Iterar sobre os elementos da coluna
    for elemento in coluna_repetidos:
        elemento_str = str(elemento)

        # Adicionar o valor ao dicionário ou somar se já existir
        if elemento_str in somas_repetidas:
            somas_repetidas[elemento_str] += elemento
        else:
            somas_repetidas[elemento_str] = elemento

    # Iterar sobre o dicionário e imprimir as repetições e somas
    for valor, soma in somas_repetidas.items():
        repeticoes = coluna_repetidos[coluna_repetidos.astype(str) == valor].count()

        if repeticoes >= 2:
            print(f"Repetição do {valor}: {repeticoes} vezes e gera a soma: {soma}")

            indices_checagem = coluna_checagem[coluna_checagem == soma].index + 2 #TODO CORRIGIR
            
            # Imprimir os índices correspondentes
            if not indices_checagem.empty:
                print(f"A soma {soma} está nas linhas {indices_checagem.to_list()} da REDE")
            else:
                print(f"A soma {soma} não foi encontrada na coluna_checagem")


def Checando_pares(coluna_REDE, coluna_w3rp):
    pares_encontrados = []
    sem_par_REDE = []
    sem_par_w3rp = []

    for i_REDE, valor_REDE in enumerate(coluna_REDE):
        encontrado = False
        start_index_w3rp = pares_encontrados[-1][1][-1] + 1 if pares_encontrados else 0

        for i_w3rp in range(start_index_w3rp, len(coluna_w3rp)):
            valor_w3rp = coluna_w3rp[i_w3rp]

            if valor_REDE == valor_w3rp:
                pares_encontrados.append((i_REDE, [i_w3rp]))
                encontrado = True
                break

        if not encontrado:
            sem_par_REDE.append((i_REDE, valor_REDE))
            sem_par_w3rp.append(None)

    print("Pares Encontrados:")
    for par in pares_encontrados:
        i_REDE, indices_w3rp = par
        valor_REDE = coluna_REDE[i_REDE]
        valores_w3rp = [coluna_w3rp[i] for i in indices_w3rp]
        print(f"({i_REDE}: {valor_REDE}, {', '.join([f'{i}:{coluna_w3rp[i]}' for i in indices_w3rp])})")

    print("\nSem Par na coluna_REDE:")
    for sem_par in sem_par_REDE:
        i_REDE, valor_REDE = sem_par
        print(f"({i_REDE}: {valor_REDE}, None)")

    print("\nSem Par na coluna_w3rp:")
    for sem_par in sem_par_w3rp:
        print(f"None")

    return pares_encontrados, sem_par_REDE, sem_par_w3rp





if __name__ == "__main__":

    var_csv = ''
    var_xlsx = ''
    root = None  # Inicializando a variável global root

    script_directory = os.path.dirname(os.path.abspath(__file__))

    # Define o diretório de trabalho atual
    os.chdir(script_directory)

    criar_janela()